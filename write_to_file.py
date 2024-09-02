import csv
from io import BytesIO

import requests
from PIL import Image as PILImage
from PIL.Image import Image
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook

from settings import METRICS


def write_to_excel(aggregated_data, file_path, include_images=True):
    """Write the aggregated data to an Excel file with optional thumbnails."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Insights"

    # Determine headers based on whether images are included
    headers = METRICS + ["media_id"]
    if include_images:
        headers = ["Thumbnail Image"] + headers

    ws.append(headers)  # Write the header row

    # Set column width and adjust text wrapping
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Customize column widths as per your needs
    if include_images:
        ws.column_dimensions['A'].width = 20  # Set Thumbnail Image column width
    ws.column_dimensions['B' if include_images else 'A'].width = 40  # Set Caption column width
    ws.column_dimensions['C' if include_images else 'B'].width = 12  # Set Timestamp column width
    ws.column_dimensions['E' if include_images else 'D'].width = 15  # Set Accounts Reached column width
    ws.column_dimensions['F' if include_images else 'E'].width = 9  # Set Comments column width
    ws.column_dimensions['G' if include_images else 'F'].width = 9  # Set Likes column width
    ws.column_dimensions['H' if include_images else 'G'].width = 6.17  # Set Saved column width
    ws.column_dimensions['I' if include_images else 'H'].width = 6.17  # Set Shares column width
    ws.column_dimensions['J' if include_images else 'I'].width = 9  # Set Initial Plays column width
    ws.column_dimensions['R' if include_images else 'Q'].width = 104  # Set Hashtags column width

    for media_id, metrics in aggregated_data.items():
        if include_images:
            # Populate the row with the relevant data including a placeholder for the image
            row = [None] + [metrics.get(metric) for metric in METRICS] + [media_id]
        else:
            # Populate the row without the image
            row = [metrics.get(metric) for metric in METRICS] + [media_id]
        ws.append(row)

        if include_images:
            # Insert the thumbnail image into the first column of the Excel sheet
            thumbnail_url = metrics.get("Thumbnail URL")
            if thumbnail_url:
                try:
                    image_response = requests.get(thumbnail_url)
                    image_response.raise_for_status()
                    img = PILImage.open(BytesIO(image_response.content))
                    img.thumbnail((100, 100))  # Resize the image to fit in the cell

                    image_io = BytesIO()
                    img.save(image_io, format="PNG")
                    image_io.seek(0)

                    img_to_insert = Image(image_io)
                    img_to_insert.anchor = f"A{ws.max_row}"  # Place the image in column A
                    img_to_insert.width, img_to_insert.height = img.size  # Set image size
                    ws.row_dimensions[ws.max_row].height = 80  # Set row height to match the thumbnail
                    ws.add_image(img_to_insert)
                except Exception as e:
                    print(f"Failed to load image {thumbnail_url}: {e}")
        else:
            # If no images, just adjust row height based on content
            ws.row_dimensions[ws.max_row].height = 20  # Adjust row height if no images

    wb.save(file_path)


def write_to_csv(aggregated_data, file_path):
    """Write the aggregated data to a CSV file."""
    headers = ["media_id"] + METRICS
    with open(file_path, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()  # Ensure header is written only once
        for media_id, metrics in aggregated_data.items():
            row = {"media_id": media_id}
            row.update(metrics)
            writer.writerow(row)
